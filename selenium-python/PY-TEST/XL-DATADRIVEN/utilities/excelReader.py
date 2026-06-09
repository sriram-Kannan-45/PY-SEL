from openpyxl import Workbook
import openpyxl


def get_data(path, sheet_name):
    final_list = []
    Workbook = openpyxl.load_workbook(path)
    sheet = Workbook[sheet_name]
    total_row = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_row + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)
    return final_list

