import os
import openpyxl

def get_data(path, sheet_name):

    base_path = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_path, path)

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook[sheet_name]

    data = []

    for r in range(2, sheet.max_row + 1):
        row = []
        for c in range(1, sheet.max_column + 1):
            row.append(sheet.cell(r, c).value)
        data.append(row)

    return data